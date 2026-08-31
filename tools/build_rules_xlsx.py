"""현재 점수 규칙을 엑셀로 뽑는다. 나중에 이 파일로 배점을 바꿀 수 있게  Cond/점수 칸을 나눈다."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "규칙_점수표_v52.xlsx"

HEADERS = [
    "번호",
    "적용단계",
    "그룹",
    "항목키",
    "항목이름",
    "조건",
    "판단기준",
    "점수",
    "배점키",
    "배점기본값",
    "비고",
]


def rows() -> list[list]:
    w = {
        "base": 10,
        "trend": 2,
        "down_line_break": 1,
        "up_line_break": -1,
        "trendline_dir_down": -2,
        "trendline_dir_up": 1,
        "trendline_up_near": 1,
        "support_near": 1,
        "support_break": -2,
        "resist_near": -1,
        "vol_sup_air": -1,
        "vol_sup_room": 1,
        "poc": 1,
        "val": 1,
        "rsi": 1,
        "ma20": -1,
        "chg1_50": -1,
        "chg1_100": -2,
        "chg1_down20": 1,
        "chg1_down30": 2,
        "chg6_100": -1,
        "chg6_300": -2,
        "chg6_400": -3,
        "rr_penalty": -1,
        "option_wall": 1,
    }
    data = [
        ["기존규칙", "시작", "base", "기본", "항상", "중립 시작점", w["base"], "base", w["base"], "SCORE_BASE. 합산 전 기본점"],
        ["기존규칙", "추세", "trend", "추세", "조회 1·2개월(lookback_days≤60) 이고 상승", "추세 추종", w["trend"], "trend", w["trend"], "배점의 절댓값 사용 후 부호 적용"],
        ["기존규칙", "추세", "trend", "추세", "조회 1·2개월 이고 하락", "추세 추종", -w["trend"], "trend", w["trend"], ""],
        ["기존규칙", "추세", "trend", "추세", "조회 1·2개월 이고 횡보", "해당 없음", 0, "trend", w["trend"], ""],
        ["기존규칙", "추세", "trend", "추세", "조회 3개월 이상 이고 상승", "고점 추격 감점", -w["trend"], "trend", w["trend"], "눌림 매수 모드"],
        ["기존규칙", "추세", "trend", "추세", "조회 3개월 이상 이고 하락", "눌림 매수 가점", w["trend"], "trend", w["trend"], ""],
        ["기존규칙", "추세", "trend", "추세", "조회 3개월 이상 이고 횡보", "해당 없음", 0, "trend", w["trend"], ""],
        ["기존규칙", "추세선 이탈", "down_line_break", "하락 추세선 이탈", "하락선 없음 / 계산 실패", "해당 없음", 0, "down_line_break", w["down_line_break"], ""],
        ["기존규칙", "추세선 이탈", "down_line_break", "하락 추세선 이탈", "하락선 위 연속 봉 수 ≥ 6", "가점 종료", 0, "down_line_break", w["down_line_break"], ""],
        ["기존규칙", "추세선 이탈", "down_line_break", "하락 추세선 이탈", "하락선 위 연속 봉 수 4~5", "이탈 유지", w["down_line_break"], "down_line_break", w["down_line_break"], ""],
        ["기존규칙", "추세선 이탈", "down_line_break", "하락 추세선 이탈", "하락선 위 연속 봉 수 1~3", "4봉 미만", 0, "down_line_break", w["down_line_break"], ""],
        ["기존규칙", "추세선 이탈", "down_line_break", "하락 추세선 이탈", "하락선 아래", "해당 없음", 0, "down_line_break", w["down_line_break"], ""],
        ["기존규칙", "추세선 이탈", "up_line_break", "상승 추세선 이탈", "상승선 없음 / 계산 실패", "해당 없음", 0, "up_line_break", w["up_line_break"], ""],
        ["기존규칙", "추세선 이탈", "up_line_break", "상승 추세선 이탈", "상승선 아래 연속 봉 수 ≥ 6", "감점 종료", 0, "up_line_break", w["up_line_break"], ""],
        ["기존규칙", "추세선 이탈", "up_line_break", "상승 추세선 이탈", "상승선 아래 연속 봉 수 4~5", "이탈 유지", w["up_line_break"], "up_line_break", w["up_line_break"], ""],
        ["기존규칙", "추세선 이탈", "up_line_break", "상승 추세선 이탈", "상승선 아래 연속 봉 수 1~3", "4봉 미만", 0, "up_line_break", w["up_line_break"], ""],
        ["기존규칙", "추세선 이탈", "up_line_break", "상승 추세선 이탈", "상승선 위", "해당 없음", 0, "up_line_break", w["up_line_break"], ""],
        ["기존규칙", "추세선 방향", "trendline_dir_down", "추세선 방향성", "하락선 이탈 가점이 있는 경우", "미적용", 0, "trendline_dir_down", w["trendline_dir_down"], "하락선 이탈 4~5봉 가점 있으면 방향성 전체 0"],
        ["기존규칙", "추세선 방향", "trendline_dir_down", "추세선 방향성", "상승선 하락 AND 하락선 하락", "둘 다 하락", w["trendline_dir_down"], "trendline_dir_down", w["trendline_dir_down"], ""],
        ["기존규칙", "추세선 방향", "trendline_dir_up", "추세선 방향성", "상승선 상승 AND 하락선 상승 AND 간격 벌어짐", "둘 다 상승·벌어짐", w["trendline_dir_up"], "trendline_dir_up", w["trendline_dir_up"], ""],
        ["기존규칙", "추세선 방향", "trendline_dir_up", "추세선 방향성", "상승선 상승 AND 하락선 상승 AND 모임", "둘 다 상승·모임", 0, "trendline_dir_up", w["trendline_dir_up"], ""],
        ["기존규칙", "추세선 방향", "trendline_dir_up", "추세선 방향성", "상승선 상승 AND 하락선 상승 AND 평행", "둘 다 상승·평행", 0, "trendline_dir_up", w["trendline_dir_up"], ""],
        ["기존규칙", "추세선 방향", "trendline_dir_up", "추세선 방향성", "상승선 상승 AND 하락선 하락", "해당 없음", 0, "trendline_dir_up", w["trendline_dir_up"], "가점/감점 없음"],
        ["기존규칙", "추세선 방향", "trendline_dir_up", "추세선 방향성", "선 없음 / 그 외 조합", "해당 없음", 0, "trendline_dir_up", w["trendline_dir_up"], ""],
        ["기존규칙", "추세선 근접", "trendline_up_near", "상승선 근접", "둘 다 상승 AND |현재가−상승선| ≤ 근처폭", "근처", w["trendline_up_near"], "trendline_up_near", w["trendline_up_near"], "근처폭 = max(ATR×0.45, 가격×0.8%)"],
        ["기존규칙", "추세선 근접", "trendline_up_near", "상승선 근접", "둘 다 상승이지만 근처가 아님", "이격", 0, "trendline_up_near", w["trendline_up_near"], ""],
        ["기존규칙", "추세선 근접", "trendline_up_near", "상승선 근접", "둘 다 상승이 아니거나 상승선 없음", "해당 없음", 0, "trendline_up_near", w["trendline_up_near"], ""],
        ["기존규칙", "지지저항", "support_near", "지지 근접", "현재가−지지 ≤ 근처폭 AND 강도 ≥ 4", "근접·강함", w["support_near"], "support_near", w["support_near"], ""],
        ["기존규칙", "지지저항", "support_near", "지지 근접", "현재가−지지 ≤ 근처폭 AND 강도 < 4", "근접·약함", 0, "support_near", w["support_near"], ""],
        ["기존규칙", "지지저항", "support_near", "지지 근접", "지지까지 이격 > 근처폭", "멀리", 0, "support_near", w["support_near"], ""],
        ["기존규칙", "지지저항", "support_break", "지지 이탈", "현재가 < 지지 − ATR×0.15", "이탈", w["support_break"], "support_break", w["support_break"], ""],
        ["기존규칙", "지지저항", "resist_near", "저항 근접", "저항−현재가 ≤ 근처폭 AND 강도 ≥ 4", "근접·강함", w["resist_near"], "resist_near", w["resist_near"], ""],
        ["기존규칙", "지지저항", "resist_near", "저항 근접", "저항−현재가 ≤ 근처폭 AND 강도 < 4", "근접·약함", 0, "resist_near", w["resist_near"], ""],
        ["기존규칙", "지지저항", "resist_near", "저항 근접", "저항까지 이격 > 근처폭", "멀리", 0, "resist_near", w["resist_near"], ""],
        ["기존규칙", "매물대", "vol_sup_air", "약한 매물대·아래 공백", "가장 가까운 매물대 지지 강도 ≥ 1", "강함", 0, "vol_sup_air", w["vol_sup_air"], "강도 1 이상이면 매물대 가감 없음"],
        ["기존규칙", "매물대", "vol_sup_air", "약한 매물대·아래 공백", "매물대 지지 강도 < 1 AND 다음 지지까지 ≥ 10%", "아래 공백", w["vol_sup_air"], "vol_sup_air", w["vol_sup_air"], ""],
        ["기존규칙", "매물대", "vol_sup_room", "약한 매물대·위 여유", "매물대 지지 강도 < 1 AND 다음 저항까지 ≥ 10% (아래 공백 조건 아닐 때)", "위 여유", w["vol_sup_room"], "vol_sup_room", w["vol_sup_room"], "아래 공백이 우선"],
        ["기존규칙", "매물대", "vol_sup_room", "약한 매물대", "매물대 지지 강도 < 1 AND 위·아래 이격 둘 다 10% 미만", "해당 없음", 0, "vol_sup_room", w["vol_sup_room"], ""],
        ["기존규칙", "밸류영역", "poc", "POC", "|현재가−POC| ≤ 근처폭 AND 상승 추세", "POC 부근 상승", w["poc"], "poc", w["poc"], "POC가 맞으면 VAL은 미적용"],
        ["기존규칙", "밸류영역", "poc", "POC", "|현재가−POC| ≤ 근처폭 AND 하락 추세", "POC 부근 하락", -w["poc"], "poc", w["poc"], ""],
        ["기존규칙", "밸류영역", "poc", "POC", "|현재가−POC| ≤ 근처폭 AND 횡보", "POC 부근 횡보", 0, "poc", w["poc"], ""],
        ["기존규칙", "밸류영역", "poc", "POC", "POC 근처가 아님", "해당 없음", 0, "poc", w["poc"], ""],
        ["기존규칙", "밸류영역", "val", "VAL", "현재가 < VAL AND 하락 추세 AND POC 근처 아님", "VAL 아래 하락", -w["val"], "val", w["val"], ""],
        ["기존규칙", "밸류영역", "val", "VAL", "현재가 < VAL AND 상승 추세 AND POC 근처 아님", "VAL 아래 상승", w["val"], "val", w["val"], ""],
        ["기존규칙", "밸류영역", "val", "VAL", "현재가 < VAL AND 횡보 AND POC 근처 아님", "VAL 아래 횡보", 0, "val", w["val"], ""],
        ["기존규칙", "밸류영역", "val", "VAL", "현재가 > VAH", "VAH 위", 0, "val", w["val"], "감점 없음"],
        ["기존규칙", "밸류영역", "val", "VAL", "VAL ≤ 현재가 ≤ VAH (밸류 내부, POC 근처 아님)", "구간 내부", 0, "val", w["val"], ""],
        ["기존규칙", "지표", "rsi", "RSI", "RSI ≥ 60", "과열", -w["rsi"], "rsi", w["rsi"], ""],
        ["기존규칙", "지표", "rsi", "RSI", "RSI ≤ 40", "과매도", w["rsi"], "rsi", w["rsi"], ""],
        ["기존규칙", "지표", "rsi", "RSI", "40 < RSI < 60", "중립", 0, "rsi", w["rsi"], ""],
        ["기존규칙", "지표", "ma20", "MA20 아래", "현재가 > MA20", "이평 위", 0, "ma20", w["ma20"], ""],
        ["기존규칙", "지표", "ma20", "MA20 아래", "현재가 < MA20 AND 추세 ≠ 상승", "이평 아래", w["ma20"], "ma20", w["ma20"], ""],
        ["기존규칙", "지표", "ma20", "MA20 아래", "현재가 < MA20 AND 상승 추세", "상승 추세라 감점 없음", 0, "ma20", w["ma20"], ""],
        ["기존규칙", "기간수익률", "chg1_100", "1개월 상승률", "30일 전 대비 ≥ 100%", "100% 이상", w["chg1_100"], "chg1_100", w["chg1_100"], ""],
        ["기존규칙", "기간수익률", "chg1_50", "1개월 상승률", "30일 전 대비 ≥ 30% AND < 100%", "30~100%", w["chg1_50"], "chg1_50", w["chg1_50"], ""],
        ["기존규칙", "기간수익률", "chg1_down30", "1개월 하락률", "30일 전 대비 ≤ −30%", "30% 이상 하락", w["chg1_down30"], "chg1_down30", w["chg1_down30"], ""],
        ["기존규칙", "기간수익률", "chg1_down20", "1개월 하락률", "30일 전 대비 ≤ −20% AND > −30%", "20~30% 하락", w["chg1_down20"], "chg1_down20", w["chg1_down20"], ""],
        ["기존규칙", "기간수익률", "chg1_50", "1개월 상승률", "그 외 (−20% ~ +30%)", "해당 없음", 0, "chg1_50", w["chg1_50"], ""],
        ["기존규칙", "기간수익률", "chg6_400", "6개월 상승률", "6개월 전 대비 ≥ 400%", "400% 이상", w["chg6_400"], "chg6_400", w["chg6_400"], "모든 조회 기간"],
        ["기존규칙", "기간수익률", "chg6_300", "6개월 상승률", "6개월 전 대비 ≥ 300% AND < 400%", "300~400%", w["chg6_300"], "chg6_300", w["chg6_300"], ""],
        ["기존규칙", "기간수익률", "chg6_100", "6개월 상승률", "6개월 전 대비 ≥ 100% AND < 300%", "100~300%", w["chg6_100"], "chg6_100", w["chg6_100"], ""],
        ["기존규칙", "기간수익률", "chg6_100", "6개월 상승률", "6개월 전 대비 < 100%", "해당 없음", 0, "chg6_100", w["chg6_100"], ""],
        ["기존규칙", "손익비", "rr_penalty", "손익비 부족", "손익비 < 1.2 AND 현재합산 ≥ 기본점+2", "여유 부족", w["rr_penalty"], "rr_penalty", w["rr_penalty"], "손절=지지−ATR×0.35, 목표=저항, RR=보상/위험"],
        ["기존규칙", "손익비", "rr_penalty", "손익비 부족", "손익비 ≥ 1.2 또는 점수 낮음", "해당 없음", 0, "rr_penalty", w["rr_penalty"], ""],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 규칙 결과가 홀딩", "미적용", 0, "option_wall", w["option_wall"], "홀딩이면 옵션 평가 안 함"],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 매도(약한/보통/강한) AND 만기≤14일 AND 근처 콜월 두껍 AND 풋월 얇음", "콜두껍/풋얇", -w["option_wall"], "option_wall", w["option_wall"], "근처=현재가±5%, 멀리=>8%, 두꺼움=상대 OI×1.8 또는 체인최대의 25%"],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 매도 AND 만기≤14일 AND 근처 풋월 두껍 AND 콜월 얇음", "풋두껍/콜얇", w["option_wall"], "option_wall", w["option_wall"], ""],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 매도 AND (콜·풋 멀리 또는 둘 다 얇음 또는 그 외)", "해당 없음", 0, "option_wall", w["option_wall"], "미국 주식 당일만"],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 매수(약한/보통/강한) AND 만기≤14일 AND 근처 풋얇 AND 근처 콜두껍", "풋얇/콜두껍", -w["option_wall"], "option_wall", w["option_wall"], ""],
        ["옵션추가", "옵션월", "option_wall", "옵션 월", "기존 매수 AND 반대이거나 그 외", "해당 없음", 0, "option_wall", w["option_wall"], "매수에서는 +1 없음"],
        ["제안컷", "합산변환", "score_pct", "합산 %", "점수 < 0", "바닥 0점 후 %", "round((점수−(−5))/(19−(−5))×100)", "", "", "SCORE_LO=-5, SCORE_HI=19. 0~100으로 자름. 0점=21%"],
        ["제안컷", "합산변환", "score_pct", "합산 %", "점수 ≥ 0", "% 변환", "round((점수+5)/24×100)", "", "", "예: 10점=63%, 19점=100%"],
        ["제안컷", "매수컷", "buy_strong", "강한 매수", "합산 % ≥ 79", "강한 매수", "", "buy_strong", 79, "기본값. 사이드바에서 변경 가능"],
        ["제안컷", "매수컷", "buy_mid", "매수", "합산 % ≥ 75 AND < 79", "매수", "", "buy_mid", 75, ""],
        ["제안컷", "매수컷", "buy_weak", "약한 매수", "합산 % ≥ 70 AND < 75", "약한 매수", "", "buy_weak", 70, ""],
        ["제안컷", "매도컷", "sell_weak", "약한 매도", "합산 % ≤ 35 AND > 30", "약한 매도", "", "sell_weak", 35, ""],
        ["제안컷", "매도컷", "sell_mid", "매도", "합산 % ≤ 30 AND > 25", "매도", "", "sell_mid", 30, ""],
        ["제안컷", "매도컷", "sell_strong", "강한 매도", "합산 % ≤ 25", "강한 매도", "", "sell_strong", 25, ""],
        ["제안컷", "홀딩", "", "홀딩", "위 매수·매도 컷에 안 들어옴", "홀딩", "", "", "", "옵션 평가 안 함"],
    ]
    out = []
    for i, row in enumerate(data, 1):
        out.append([i] + row)
    return out


def params() -> list[list]:
    return [
        ["이름", "값", "설명"],
        ["규칙버전", 52, "SIGNAL_RULE_VERSION"],
        ["기본점 SCORE_BASE", 10, "중립 시작"],
        ["합산 최저 SCORE_LO", -5, "10−15. % 눈금 하한"],
        ["합산 최고 SCORE_HI", 19, "10+9. % 눈금 상한"],
        ["근처폭", "max(ATR×0.45, 가격×0.8%)", "지지/저항/상승선/POC 근처 판정"],
        ["지지 이탈 여유", "ATR×0.15", "현재가가 지지보다 이만큼 더 아래면 이탈"],
        ["손절", "지지 − ATR×0.35", "손익비 계산"],
        ["1차 목표", "가장 가까운 저항", "손익비 계산"],
        ["손익비 기준", 1.2, "미만이고 점수가 기본+2 이상이면 감점"],
        ["추세 추종 조회", "lookback_days ≤ 60", "1·2개월. 그 이상은 눌림 매수"],
        ["RSI 과매도", 40, "이하 +"],
        ["RSI 과열", 60, "이상 −"],
        ["지지/저항 강도 기준", 4, "4 미만이면 근접해도 가감 없음"],
        ["매물대 약함", 1, "강도 1 미만일 때만 공백/여유 규칙"],
        ["매물대 이격", "10%", "다음 지지 또는 저항까지"],
        ["옵션 만기", "14일 이내", "그 이상이면 0점"],
        ["옵션 근처", "현재가 ±5%", "OPTION_NEAR_PCT=0.05"],
        ["옵션 멀리", "현재가 ±8% 초과", "OPTION_FAR_PCT=0.08"],
        ["옵션 두꺼움 비율", "상대 OI × 1.8 또는 체인 최대 OI의 25%", ""],
        ["옵션 얇음", "OI=0 또는 체인 최대의 12% 미만 또는 상대의 1/1.8", ""],
        ["옵션 적용 시장", "미국 주식 당일(시차 1일 안)", "한국·코인·과거·홀딩은 미적용"],
        ["합산 후 처리", "음수면 0으로 자름", "max(0, score) 후 %"],
    ]


def style_header(ws, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for col in range(1, n_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    align = Alignment(vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align
            cell.border = thin


def main() -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide["A1"] = "자산 트레이드 분석기 점수 규칙표 (v52)"
    guide["A1"].font = Font(size=16, bold=True, color="1F4E79")
    guide.merge_cells("A1:B1")
    notes = [
        "",
        "이 파일은 지금 코드에 들어 있는 규칙을 풀어 적은 표입니다.",
        "다음에 점수를 어떻게 매길지 알려주실 때, 「규칙항목」 시트의 「점수」 칸을 고치거나 행을 추가해 주세요.",
        "",
        "시트",
        "안내 — 이 설명",
        "규칙항목 — 조건별 점수. 한 행이 하나의 판단입니다.",
        "공통파라미터 — 근처폭, ATR 배수, 옵션 비율 등 숫자.",
        "",
        "점수 매기는 순서",
        "1) 「기존규칙」 행을 모두 더합니다. 음수면 0으로 자릅니다.",
        "2) (점수 + 5) / 24 × 100 을 반올림해 합산 %를 만듭니다. (범위 −5~19점)",
        "3) 합산 %로 약한매수/매수/강한매수/약한매도/매도/강한매도/홀딩을 정합니다.",
        "4) 홀딩이면 옵션을 넣지 않습니다. 매수·매도면 「옵션추가」 점수를 더하고 %와 제안을 다시 계산합니다.",
        "",
        "열 설명 (규칙항목)",
        "적용단계 — 기존규칙 / 옵션추가 / 제안컷",
        "항목키 — 코드 식별자. 바꾸지 않는 것이 좋습니다.",
        "조건 — 언제 이 행이 쓰이는지",
        "판단기준 — 짧은 이름",
        "점수 — 실제로 더하는 점수. 이 칸을 고치면 됩니다.",
        "배점키 / 배점기본값 — 사이드바에서 크기를 바꿀 수 있는 항목",
    ]
    for i, line in enumerate(notes, 2):
        guide[f"A{i}"] = line
        if line.startswith("시트") or line.startswith("점수 매기는") or line.startswith("열 설명"):
            guide[f"A{i}"].font = Font(bold=True, color="1F4E79")
    guide.column_dimensions["A"].width = 88

    ws = wb.create_sheet("규칙항목")
    ws.append(HEADERS)
    yellow = PatternFill("solid", fgColor="FFF2CC")
    for row in rows():
        ws.append(row)
        ws.cell(ws.max_row, 8).fill = yellow
    style_header(ws, len(HEADERS))
    autosize(ws, {1: 8, 2: 12, 3: 14, 4: 18, 5: 18, 6: 62, 7: 22, 8: 12, 9: 18, 10: 12, 11: 42})
    ws["H1"].fill = PatternFill("solid", fgColor="C65911")
    ws.row_dimensions[1].height = 24

    wp = wb.create_sheet("공통파라미터")
    for row in params():
        wp.append(row)
    style_header(wp, 3)
    autosize(wp, {1: 22, 2: 42, 3: 70})
    for r in range(2, wp.max_row + 1):
        wp.cell(r, 2).fill = PatternFill("solid", fgColor="E2EFDA")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
